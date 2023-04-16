import requests
from sys import argv
from bs4 import BeautifulSoup
import urllib
import re
from openpyxl import load_workbook

def replace_chars(input_string):
    input_string = input_string.replace("]", "")
    input_string = input_string.replace("[", "")
    input_string = input_string.replace("'", "")
    input_string = input_string.replace("\\xa0", "")
    input_string = input_string.replace("\\n", "")
    return input_string

class NoFluffJobs():
    # URL = argv[1]
    def __init__(self):
        self.jobs_dict = {}
    
    def getDomainName(self, url):
        try:
            parsed_uri = urllib.request.urlparse(url)
            domainName = '{uri.scheme}://{uri.netloc}'.format(uri=parsed_uri)
            return domainName
        except Exception as e:
            print(f"Exception {e} on getDomainName.")

    def getPagesCount(self, url):
        try:
            page = requests.get(url)
            soup = BeautifulSoup(page.content, "html.parser")
            pages_count = soup.find_all('a', {'class': 'page-link'})
            max_page_count = 1
            for page in pages_count:
                try:
                    val = int(page.text.strip())
                except(ValueError):
                    continue
                max_page_count = val if val > max_page_count else max_page_count
            print(f"All found pages: {max_page_count}")
            return max_page_count
        except Exception as e:
            print(f"Exception {e} on getPagesCount.")        

    def updateJobsDict(self, url):
        max_pages = self.getPagesCount(url)
        domainName = self.getDomainName(url)
        try:
            for page_num in range(1, max_pages+1):
                page = requests.get(url+f"&page={page_num}")
                # print(page_num)
                page_soup = BeautifulSoup(page.content, "html.parser")
                
                job_links_list = page_soup.find_all("a", {"class": "posting-list-item"})

                for job in job_links_list:
                    job_link = domainName+job['href']
                    job_title = job.find('h3').text
                    job_company = job.find('span', class_=re.compile("company", re.I)).text 
                    job_salary = job.find('span', class_=re.compile("badgy salary", re.I)).text
                    job_location = job.find('div', class_=re.compile("tw-flex tw-items-center ng-star-inserted", re.I)).text
                    
                    self.jobs_dict[job_link] = {"Title": [], "Company": [], "Salary": [], "Location": []}
                    self.jobs_dict[job_link]["Title"].append(job_title)
                    self.jobs_dict[job_link]["Company"].append(job_company)
                    self.jobs_dict[job_link]["Salary"].append(job_salary)
                    self.jobs_dict[job_link]["Location"].append(job_location)
                    #print(len(self.jobs_dict))
        # print(self.jobs_dict)
        except Exception as e:
            print(f"Exception {e} on updateJobsDict.")                
    
    def updateExcel(self):
        try:
            workbook = load_workbook("jobs.xlsx")
            # sheet = workbook.active
            sheet = workbook['NoFluff']
            for k, v in self.jobs_dict.items():
                exists = False
                for row in sheet.rows:
                    if row[0].value is not None and k in row[0].value:
                        # print(f"Already in excel: {k}")
                        exists = True
                if exists is False:
                    # print(f"Fresh one: {k}")
                    sheet.insert_rows(2, 1)
                    sheet.cell(row = 2, column = 1, value = '=HYPERLINK("{}", "{}")'.format(k, f"{k}"))
                    sheet.cell(row = 2, column = 2, value = replace_chars(str(v["Title"])))
                    sheet.cell(row = 2, column = 3, value = replace_chars(str(v["Company"])))
                    sheet.cell(row = 2, column = 4, value = replace_chars(str(v["Salary"])))
                    sheet.cell(row = 2, column = 5, value = replace_chars(str(v["Salary"])))
            workbook.save(filename="jobs.xlsx")
        except Exception as e:
            print(f"Exception: {e} on updateExcel.")
        
    
fluff = NoFluffJobs()
fluff.updateJobsDict("https://nofluffjobs.com/pl/testing?criteria=employment%3Db2b%20requirement%3DPython")
fluff.updateExcel()