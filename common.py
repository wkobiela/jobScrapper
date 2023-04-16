import urllib
from openpyxl import load_workbook

def replace_chars(input_string):
    input_string = input_string.replace("]", "")
    input_string = input_string.replace("[", "")
    input_string = input_string.replace("'", "")
    input_string = input_string.replace("\\xa0", "")
    input_string = input_string.replace("\\n", "")
    return input_string

def getDomainName(url):
    try:
        parsed_uri = urllib.request.urlparse(url)
        domainName = '{uri.scheme}://{uri.netloc}'.format(uri=parsed_uri)
        return domainName
    except Exception as e:
        print(f"Exception {e} on getDomainName.")
        
def updateExcel(sheet, jobs_dict):
    try:
        workbook = load_workbook("jobs.xlsx")
        # sheet = workbook.active
        sheet = workbook[f"{sheet}"]
        for k, v in jobs_dict.items():
            exists = False
            for row in sheet.rows:
                if row[0].value is not None and k in row[0].value:
                    # print(f"Already in excel: {k}")
                    exists = True
            if exists is False:
                # print(f"Fresh one: {k}")
                sheet.insert_rows(2, 1)
                sheet.cell(row = 2, column = 1, value = '=HYPERLINK("{}", "{}")'.format(k, f"{k}"))
                sheet.cell(row = 2, column = 2, value = replace_chars(str(v["Title"])))
                sheet.cell(row = 2, column = 3, value = replace_chars(str(v["Company"])))
                sheet.cell(row = 2, column = 4, value = replace_chars(str(v["Salary"])))
                sheet.cell(row = 2, column = 5, value = replace_chars(str(v["Location"])))
        workbook.save(filename="jobs.xlsx")
    except Exception as e:
        print(f"Exception: {e} on updateExcel.")