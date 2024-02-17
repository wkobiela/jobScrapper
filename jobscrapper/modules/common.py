import re
import os
import sys
import logging
import urllib
from datetime import datetime
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from unidecode import unidecode

now = datetime.now()

log = logging.getLogger(__name__)

def checkFileExistance(filename):
    path = f"{os.getcwd()}/{filename}"
    check_file = os.path.isfile(path)
    return check_file

def replaceChars(input_string):
    input_string = input_string.replace("]", "")
    input_string = input_string.replace("[", "")
    input_string = input_string.replace("'", "")
    input_string = input_string.replace("\\xa0", "")
    input_string = input_string.replace("\\n", "")
    return input_string

def getDomainName(url):
    try:
        given_url = urllib.parse.urlparse(url)
        domainName = given_url.netloc
        
        return domainName
    except Exception as e:
        log.error('common:getDomainName: Exception: %s.', e)
        raise

def getPagesCount(url, parent, child, regex):
    try:
        page = requests.get(url, timeout=120)
        soup = BeautifulSoup(page.content, "html.parser")
        pages_count = soup.find_all(parent, {child: re.compile(regex)})
        max_page_count = 1
        for page in pages_count:
            try:
                val = int(page.text.strip())
            except(ValueError):
                continue
            max_page_count = val if val > max_page_count else max_page_count
        log.debug('common:getPagesCount: Found %s pages with offers. Scrapping further.', max_page_count)
        return max_page_count
    except Exception as e:
        log.error('common:getPagesCount: Exception: %s.', e)
        return 1
        
def updateExcel(sheet, jobs_dict):
    try:
        workbook = load_workbook("jobs.xlsx")
        new_jobs = 0
        sheet = workbook[f"{sheet}"]
        for k, v in jobs_dict.items():
            exists = False
            for row in sheet.rows:
                if row[0].value is not None and k in row[0].value:
                    exists = True
            if exists is False:
                new_jobs += 1
                sheet.insert_rows(2, 1)
                # sheet.cell(row = 2, column = 1, value = '=HYPERLINK("{}", "{}")'.format(k, f"{k}"))
                sheet.cell(row = 2, column = 1, value = f'=HYPERLINK("{k}", "{k}")')
                sheet.cell(row = 2, column = 2, value = replaceChars(str(v["Title"])))
                sheet.cell(row = 2, column = 3, value = replaceChars(str(v["Company"])))
                sheet.cell(row = 2, column = 4, value = replaceChars(str(v["Salary"])))
                sheet.cell(row = 2, column = 5, value = replaceChars(str(v["Location"])))
                sheet.cell(row = 2, column = 6, value = now.strftime("%d/%m/%Y, %H:%M"))
        if new_jobs > 0:
            log.info('common:updateExcel: %s new offers in %s!', new_jobs, sheet.title)
        else:
            log.info('common:updateExcel: No new offers in %s.', sheet.title)
        workbook.save(filename="jobs.xlsx")
    except Exception as e:
        log.error('common:updateExcel: Exception: %s', e)
        
def createLinks(**kwargs):
    if not all(key in kwargs for key in ('site','role','lvl','city')):
        log.error("common:createLinks: Not enough or too much arguments. \
            Please fill only and all of the following: site, role, lvl, city.")
        sys.exit()
        
    for key, item in kwargs.items():
        if key == "site":
            site = item
        elif key == "role":
            role = item
        elif key == "lvl":
            lvl = item
        elif key == "city":
            city = item
    
    if site == "BulldogJob":
        generated_link = f"https://bulldogjob.pl/companies/jobs/s/role,{role}/experienceLevel,{lvl}/city,{unidecode(city)}"
    elif site == "NoFluffJobs":
        generated_link = f"https://nofluffjobs.com/pl/praca-zdalna/{role}?criteria=city%3D{unidecode(city)}%20%20seniority%3D{lvl}"
    elif site == "JustjoinIt":
        generated_link = f"https://justjoin.it/{unidecode(city).lower()}/{role}/experience-level_{lvl}/remote_yes"
    log.debug("common:createLinks: Generated link: %s", generated_link)    
    return(generated_link)